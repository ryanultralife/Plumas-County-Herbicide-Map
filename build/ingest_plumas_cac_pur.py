#!/usr/bin/env python3
"""Ingest Plumas County's own CalAgPermits PUR export (CPRA response from the
Plumas-Sierra Ag Dept, "Plumas PUR from 01-01-24.xlsx", received 2026-07-22).

WHY a separate source: the statewide DPR extract only carries Plumas through
2024. This county export runs 2024-01-01 -> current and is the ONLY source we
have for Plumas 2025-2026. We load *only* 2025+2026 so we never double-count the
2024 applications the DPR extract already holds (the two systems use different
ids, so a plain dedup could not catch the overlap). Rows are tagged
source='pur-cac-plumas' and app_id 'purcac:{document}:{comtrs}' so the whole
batch is identifiable and reversible.

Grain matches public.applications (one row per application event = one location +
its primary herbicide's active-ingredient pounds):
  app_id            = purcac:{Document#}:{COMTRS}
  owner             = 32{yy}{permit7}      (synthetic GROWER_ID; name upserted)
  product           = primary line's Product Name
  active_ingredient = CHEMNAME from the DPR reg_no map     unit = lbs
  amount            = quantity * (lbs-AI per unit) learned from DPR's own numbers
  acres             = primary line's Treated Amount (ACRES)
  lat/lon           = CDPR Plumas PLSS section centroid (CEN_LAT84/LONG84)
  land_type         = classified from the Permitee (landowner) name
  date              = Application Date -> MM/DD/YYYY

Pounds: the county export reports raw volume (gal/oz/lb), NOT DPR-computed
pounds. build/_regmap.json gives, per EPA reg_no, the primary active ingredient
and a lbs-of-AI-per-unit rate = sum(LBS_CHM_USED)/sum(AMT_PRD_USED) taken from
the 2020-2024 DPR extract itself -> a grounded conversion, no fabrication.
Adjuvants/surfactants (spray oils, MSO) carry no AI pounds and never determine a
dot; when an event's only product is an adjuvant, active_ingredient is blank.

Derived lookups (committed, so this reruns without the 108MB DPR csv or the
shapefile):
  data/incoming/2026-07/plumas-pur/_regmap.json          (reg_no -> AI + rates)
  data/incoming/2026-07/plumas-pur/_plumas_centroids.json (COMTRS -> lat,lon)

Usage:
  python build/ingest_plumas_cac_pur.py            # transform + validate (no DB)
  DBURL=... python build/ingest_plumas_cac_pur.py --load   # + insert + names
"""
import sys, os, csv, json, time, datetime, subprocess, tempfile, collections

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "data", "incoming", "2026-07", "plumas-pur")
XLSX = os.path.join(BASE, "Plumas PUR from 01-01-24.xlsx")
REGMAP = os.path.join(BASE, "_regmap.json")
CENTROIDS = os.path.join(BASE, "_plumas_centroids.json")
OUT = os.path.join(BASE, "applications_plumas_cac_2025_2026.csv")
NAMES_OUT = os.path.join(BASE, "operator_names_plumas_cac.csv")
YEARS = {2025, 2026}

COLS = ["app_id", "source", "region", "date", "year", "lat", "lon", "county",
        "land_type", "owner", "product", "active_ingredient", "amount", "unit",
        "acres", "method", "activity", "project", "status", "url", "pulled"]

# county Quantity Units -> DPR UNIT_OF_MEAS code
UNIT = {"GALLON": "GA", "OUNCE": "OZ", "POUNDS": "LB", "POUND": "LB",
        "PINT": "PT", "QUART": "QT"}
# liquid unit -> gallons, for converting when a product lacks that unit's rate
TO_GAL = {"GA": 1.0, "QT": 0.25, "PT": 0.125, "OZ": 1.0 / 128.0}
# 2 reg_nos absent from 2020-2024 DPR, resolved by label:
HARDCODE = {
    "84229-32-ZA": {"chemname": "HEXAZINONE", "rates": {"LB": 0.75}, "adjuvant": False},
    "92617-50002-AA": {"chemname": "", "rates": {}, "adjuvant": True},  # INAPRO H: adjuvant, never sole
}


def land_type(permitee):
    p = (permitee or "").upper()
    if "USDA" in p or "NATIONAL FOREST" in p or "FOREST SERVICE" in p:
        return "federal"
    if any(k in p for k in ("COLLINS PINE", "SIERRA PACIFIC", "BEATY", "INGERSOLL",
                            "TIMBER", "FOREST", "RESOURCE CON")):
        return "forestry"
    return "ag"  # ranches, college ag farm, Russell Reid, etc.


def yr(v):
    if isinstance(v, datetime.datetime):
        return v.year
    if v is None:
        return None
    for tok in str(v).replace('/', ' ').replace('-', ' ').split():
        if len(tok) == 4 and tok.isdigit():
            return int(tok)
    return None


def fmt_date(v):
    if isinstance(v, datetime.datetime):
        return f"{v.month:02d}/{v.day:02d}/{v.year}"
    s = str(v or "").strip()
    if len(s) >= 10 and s[4] == '-':          # 2025-10-24 00:00:00
        return f"{s[5:7]}/{s[8:10]}/{s[:4]}"
    if '/' in s:                               # 4/24/2026
        m, d, y = (s.split()[0].split('/') + ['', '', ''])[:3]
        if y:
            return f"{int(m):02d}/{int(d):02d}/{y}"
    return None


def to_f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def sec2(s):
    s = str(s or "").strip()
    return s.zfill(2) if s.isdigit() else s


def lbs_per_unit(reg, unit_code):
    """lbs of active ingredient per one <unit_code> of this product, from DPR."""
    e = REG.get(reg)
    if not e or e.get("adjuvant") or not e.get("chemname"):
        return 0.0
    rates = e.get("rates", {})
    if unit_code in rates:
        return rates[unit_code]
    # fall back through gallons for liquid units
    if unit_code in TO_GAL and "GA" in rates:
        return TO_GAL[unit_code] * rates["GA"]
    pr = e.get("prod_rates", {})
    if unit_code in pr:
        return pr[unit_code]
    if unit_code in TO_GAL and "GA" in pr:
        return TO_GAL[unit_code] * pr["GA"]
    return 0.0


def chemname(reg):
    e = REG.get(reg)
    return (e or {}).get("chemname", "") if e and not (e or {}).get("adjuvant") else ""


def read_rows():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    def sheet(name, dcol):
        ws = wb[name]
        H = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
        i = {h: H.index(h) for h in H}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or r[0] is None:
                continue
            twp = str(r[i['Township']] or '').strip()
            rng = str(r[i['Range']] or '').strip()
            sec = sec2(r[i['Section']])
            mer = str(r[i['Meridian']] or 'M').strip() or 'M'
            if not (twp and rng and sec):
                continue
            yield dict(
                document=str(r[i['Document#']] or '').strip(),
                permit=str(r[i['Permit #']] or '').strip(),
                permitee=str(r[i['Permitee']] or '').strip(),
                comtrs=f"32{mer}{twp}{rng}{sec}",
                date=r[i[dcol]], year=yr(r[i[dcol]]),
                reg_no=str(r[i['EPA Reg No']] or '').strip(),
                product=str(r[i['Product Name']] or '').strip(),
                qty=to_f(r[i['Quantity Used']]),
                qty_unit=str(r[i['Quantity Units']] or '').strip().upper(),
                treated=to_f(r[i['Treated Amount']]),
                treated_unit=str(r[i['Treated Units']] or '').strip().upper(),
                applicator=(str(r[i['Applicator Name']] or '').strip() if 'Applicator Name' in i else ''),
            )
    rows = list(sheet("Single Job PURs", "Application Date")) + \
        list(sheet("Monthly Ag PURs", "Start Application Date"))
    wb.close()
    return [r for r in rows if r['year'] in YEARS]


def transform():
    rows = read_rows()
    events = collections.defaultdict(list)
    for r in rows:
        events[(r['document'], r['comtrs'])].append(r)

    recs, names = [], {}
    miss_cent = set()
    for (doc, comtrs), lines in events.items():
        # per line: computed lbs of AI + whether it's a real herbicide
        for ln in lines:
            uc = UNIT.get(ln['qty_unit'], ln['qty_unit'])
            ln['_uc'] = uc
            ln['_lbs'] = round((ln['qty'] or 0.0) * lbs_per_unit(ln['reg_no'], uc), 4)
            ln['_ai'] = chemname(ln['reg_no'])
        herb = [ln for ln in lines if ln['_ai']]
        rep = max(herb, key=lambda x: x['_lbs']) if herb else max(lines, key=lambda x: (x['_lbs'], x['qty'] or 0))
        cent = CENT.get(comtrs)
        if not cent:
            miss_cent.add(comtrs)
            continue
        year = rep['year']
        yy = str(year)[2:]
        permit7 = ''.join(ch for ch in rep['permit'] if ch.isdigit()).zfill(7)[-7:]
        owner = f"32{yy}{permit7}"
        acres = rep['treated'] if (rep['treated_unit'].startswith("ACRE")) else None
        recs.append({
            "app_id": f"purcac:{doc}:{comtrs}", "source": "pur-cac-plumas",
            "region": "northern-sierra", "date": fmt_date(rep['date']), "year": year,
            "lat": cent[0], "lon": cent[1], "county": "Plumas",
            "land_type": land_type(rep['permitee']), "owner": owner,
            "product": rep['product'], "active_ingredient": rep['_ai'],
            "amount": (rep['_lbs'] if rep['_lbs'] else None), "unit": "lbs",
            "acres": acres, "method": None, "activity": None, "project": None,
            "status": "completed", "url": "https://calpip.cdpr.ca.gov/",
            "pulled": "2026-07-22"})
        if rep['permitee']:
            names[owner] = rep['permitee']

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(recs)
    with open(NAMES_OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["operator_id", "name", "county", "source"])
        for oid, nm in sorted(names.items()):
            w.writerow([oid, nm, "Plumas", "cac-plumas-pur-2026"])

    # validation
    byyear = collections.Counter(r['year'] for r in recs)
    tot_lbs = sum(r['amount'] or 0 for r in recs)
    tot_ac = sum(r['acres'] or 0 for r in recs)
    print(f"events(single+monthly, 2025-26): {len(events):,}  ->  mappable app rows: {len(recs):,}")
    print("by year:", dict(sorted(byyear.items())))
    print(f"total AI lbs: {tot_lbs:,.0f}   total acres: {tot_ac:,.0f}")
    n_ai = sum(1 for r in recs if r['active_ingredient'])
    print(f"rows with an active ingredient (herbicide): {n_ai:,}  | adjuvant-only: {len(recs)-n_ai:,}")
    print(f"distinct owners (landowner-years) named: {len(names)}")
    if miss_cent:
        print("WARN missing centroids:", sorted(miss_cent))
    print("->", OUT)
    print("->", NAMES_OUT)
    return len(recs)


LOAD_SQL = """
set statement_timeout=0;
set client_min_messages=warning;
create temp table _pcac (like public.applications including defaults);
\\copy _pcac ({cols}) from '{csv}' csv header
select count(*) staged from _pcac;
insert into public.applications ({cols})
  select {cols} from _pcac
on conflict (app_id) do nothing;
select count(*) filter (where year=2025) y2025, count(*) filter (where year=2026) y2026,
       round(sum(amount) filter (where unit='lbs')::numeric,0) lbs,
       round(sum(acres)::numeric,0) acres
from public.applications where source='pur-cac-plumas';
create temp table _nm (operator_id text, name text, county text, source text);
\\copy _nm (operator_id,name,county,source) from '{names}' csv header
insert into public.operator_names (operator_id,name,county,source,updated)
  select operator_id,name,county,source,'2026-07-22' from _nm
on conflict (operator_id) do update
  set name=excluded.name, county=coalesce(public.operator_names.county,excluded.county),
      source=excluded.source, updated=excluded.updated;
select count(*) plumas_names_2025_26 from public.operator_names where source='cac-plumas-pur-2026';
"""


def load():
    dburl = os.environ.get("DBURL") or sys.exit("Set DBURL to load.")
    cols = ",".join(COLS)
    sql = LOAD_SQL.format(cols=cols, csv=OUT.replace(os.sep, "/"),
                          names=NAMES_OUT.replace(os.sep, "/"))
    sqlf = tempfile.NamedTemporaryFile("w", delete=False, suffix=".sql", encoding="utf-8")
    sqlf.write(sql)
    sqlf.close()
    r = subprocess.run(["psql", dburl, "-v", "ON_ERROR_STOP=1", "-f", sqlf.name],
                       capture_output=True, text=True)
    os.unlink(sqlf.name)
    print(r.stdout.strip())
    if r.returncode != 0:
        sys.exit(r.stderr.strip()[-2000:])


if __name__ == "__main__":
    # load derived lookups
    REG = json.load(open(REGMAP, encoding="utf-8"))
    for k, v in HARDCODE.items():
        REG.setdefault(k, {})
        REG[k].update({"found": True, "chemname": v["chemname"],
                       "rates": v["rates"], "adjuvant": v["adjuvant"]})
    CENT = json.load(open(CENTROIDS, encoding="utf-8"))
    globals()["REG"] = REG
    globals()["CENT"] = CENT
    n = transform()
    if "--load" in sys.argv and n:
        load()
