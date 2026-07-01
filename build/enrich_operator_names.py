#!/usr/bin/env python3
"""
Identify real operator names for CDPR PUR GROWER_IDs using the public
permit-holder datasets that some County Agricultural Commissioners publish.

Key insight (verified against our data): a PUR GROWER_ID like 27222700348 is
[reporting county 2][permit year 2][home-county+permit 7]. The trailing 7 chars
(2700348) ARE the county "permit number" published in those datasets, with the
leading 2 digits = the operator's home county. So:  right(owner,7) == PermNum.

This fetches each available county dataset -> {permnum: name}, writes a combined
CSV, then (in SQL) joins right(owner,7)=permnum across ALL our GROWER_IDs and
upserts (grower_id -> name) into public.operator_names. After that, the frontend's
loadOperatorNames() shows the real name everywhere (map popups, Data & Trends,
Source Data), overriding the coded ID.

Usage:  DBURL="postgres://..." python build/enrich_operator_names.py
"""
import os, sys, csv, io, json, re, zipfile, urllib.request, urllib.parse, subprocess, tempfile

UA = {"User-Agent": "Mozilla/5.0 (SprayMap transparency project)"}

def fetch(url, timeout=120):
    req = urllib.request.Request(url, headers=UA)
    return urllib.request.urlopen(req, timeout=timeout).read()

def norm_perm(v):
    # Produce the 7-char permit key that equals right(owner,7). Most counties use
    # a numeric permit; Contra Costa's carry a trailing permit-type letter
    # (e.g. 070125C), and GROWER_IDs preserve that letter, so keep [0-9A-Z].
    s = str(v).strip().upper().split(".")[0]   # handle "3200530.0" float-style ids
    s = re.sub(r"[^0-9A-Z]", "", s)
    if not s:
        return None
    return s[-7:] if len(s) >= 7 else s.zfill(7)

def clean_name(n):
    n = " ".join(str(n or "").split()).strip().strip(",").upper()
    return n if n and n not in ("NULL", "NONE", "N/A") else None

rows = {}   # permnum -> (name, county, source)   first non-empty wins
def add(permnum, name, county, source):
    p, nm = norm_perm(permnum), clean_name(name)
    if p and nm and p not in rows:
        rows[p] = (nm, county, source)

# ---------- generic ArcGIS paginator (FeatureServer or MapServer /query) ----------
def arcgis(base, pfield, nfield, county, source, where="1=1"):
    """base = a layer URL ending in /FeatureServer/<n> or /MapServer/<n> (no /query).
    Pages by objectIds (robust across hosted/enterprise/MapServer); falls back to
    resultOffset paging if the service won't return an id list."""
    fields = pfield + "," + nfield
    qbase = base + "/query?where=" + urllib.parse.quote(where)
    try:
        idj = json.loads(fetch(qbase + "&returnIdsOnly=true&f=json"))
        oids = idj.get("objectIds") or []
    except Exception:
        oids = []
    n = 0
    if oids:
        for i in range(0, len(oids), 250):
            chunk = oids[i:i + 250]
            d = json.loads(fetch(base + "/query?objectIds=" + ",".join(str(x) for x in chunk) +
                                 "&outFields=" + urllib.parse.quote(fields) +
                                 "&returnGeometry=false&f=json"))
            for f in d.get("features", []):
                a = f.get("attributes", {})
                add(a.get(pfield), a.get(nfield), county, source); n += 1
        return n
    off = 0                                            # fallback: offset paging
    while True:
        d = json.loads(fetch(qbase + "&outFields=" + urllib.parse.quote(fields) +
                             "&returnGeometry=false&f=json&resultRecordCount=2000&resultOffset=" + str(off)))
        feats = d.get("features", [])
        if not feats:
            break
        for f in feats:
            a = f.get("attributes", {})
            add(a.get(pfield), a.get(nfield), county, source); n += 1
        off += len(feats)
        if not d.get("exceededTransferLimit") and len(feats) < 2000:
            break
    return n

# ---------- Monterey (27): ArcGIS FeatureServer ----------
def monterey():
    base = ("https://services2.arcgis.com/nOGTdfb4kF4dZljH/arcgis/rest/services/"
            "2020RanchMapAtlasDataOD/FeatureServer/0/query")
    off, got = 0, 0
    while True:
        url = (base + "?where=1%3D1&outFields=PermNum,Permittee&returnGeometry=false"
               "&f=json&resultRecordCount=2000&resultOffset=" + str(off))
        d = json.loads(fetch(url))
        feats = d.get("features", [])
        if not feats:
            break
        for f in feats:
            a = f.get("attributes", {})
            add(a.get("PermNum"), a.get("Permittee"), "Monterey", "monterey-ranch-atlas-2020")
        got += len(feats); off += len(feats)
        if not d.get("exceededTransferLimit") and len(feats) < 2000:
            break
    return got

# ---------- Kern (15): zipped CSV ----------
def kern():
    import openpyxl
    z = zipfile.ZipFile(io.BytesIO(fetch("http://www.kernag.com/ep/permit-use/20_29/2024_kern_permit.zip")))
    xn = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
    if not xn:
        return 0
    wb = openpyxl.load_workbook(io.BytesIO(z.read(xn[0])), read_only=True, data_only=True)
    ws = wb.active
    hdr = None; pi = oi = None; n = 0
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(c or "").strip() for c in row]
            for i, h in enumerate(hdr):
                if h == "Permit Number": pi = i
                if h == "Permittee": oi = i
            if pi is None or oi is None:
                hdr = None
            continue
        if pi < len(row) and oi < len(row):
            add(row[pi], row[oi], "Kern", "kern-permits-2024"); n += 1
    return n

# ---------- Stanislaus (50): ArcGIS Hub open data (CSV download) ----------
def stanislaus():
    base = ("https://services.arcgis.com/EeYBJFxLdUojipYa/arcgis/rest/services/"
            "Permit_Sites_and_Commodities_Open_Data/FeatureServer/7/query")
    where = urllib.parse.quote("Permit_Type IN ('Op-Id','RMP')")
    off, n = 0, 0
    while True:
        url = (base + "?where=" + where + "&outFields=Permit_Number,Operator&returnGeometry=false"
               "&f=json&resultRecordCount=2000&resultOffset=" + str(off))
        d = json.loads(fetch(url))
        feats = d.get("features", [])
        if not feats:
            break
        for f in feats:
            a = f.get("attributes", {})
            add(a.get("Permit_Number"), a.get("Operator"), "Stanislaus", "stanislaus-permits")
        n += len(feats); off += len(feats)
        if not d.get("exceededTransferLimit") and len(feats) < 2000:
            break
    return n

# ---------- San Joaquin (39): XLSX ----------
def san_joaquin():
    try:
        import openpyxl
    except ImportError:
        print("  (san_joaquin skipped: openpyxl not installed)")
        return 0
    raw = fetch("https://www.sjgov.org/docs/default-source/agricultural-commissioner-documents/"
                "pur-business-lic/permits/permits-2024-.xlsx?sfvrsn=8d6c9a21_6")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    hdr = None; pi = oi = None; n = 0
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(c or "").strip() for c in row]
            for i, h in enumerate(hdr):
                if h == "Permit Number": pi = i
                if h == "Operator": oi = i
            if pi is None or oi is None:
                hdr = None
            continue
        if pi < len(row) and oi < len(row):
            add(row[pi], row[oi], "San Joaquin", "sanjoaquin-permits-2024")
            n += 1
    return n

# ---------- Plumas (32) + others already obtained: local records-request CSVs ----------
def plumas_local():
    n = 0
    for fn, pcol, ncol in [("data/single_job_pur.csv", "Permit #", "Permitee"),
                           ("data/prodag_monthly_summary.csv", "Permit #", "Permitee"),
                           ("data/nonprod_ag_mspur.csv", "Permit #", "Permittee")]:
        try:
            with open(fn, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    add(row.get(pcol), row.get(ncol), "Plumas", "plumas-records-2021-2024")
                    n += 1
        except FileNotFoundError:
            pass
    return n

# ---------- shared xlsx permit->name scanner ----------
def _xlsx_permit_name(raw, pcol, ncol, county, source, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    hdr = None; pi = oi = None; n = 0
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(c or "").strip() for c in row]
            if pcol in hdr and ncol in hdr:
                pi = hdr.index(pcol); oi = hdr.index(ncol)
            else:
                hdr = None
            continue
        if pi < len(row) and oi < len(row):
            add(row[pi], row[oi], county, source); n += 1
    return n

# ---------- Contra Costa (07): DocumentCenter XLSX (permits carry trailing letter) ----------
def contra_costa():
    raw = fetch("https://www.contracosta.ca.gov/DocumentCenter/View/91252/"
                "2026-Permittees-and-Operators-in-Contra-Costa-as-of-April-2026")
    return _xlsx_permit_name(raw, "Permit Number", "Operator", "Contra Costa",
                             "contracosta-permits-2026", sheet="PermitSiteCommSearchResults")

# ---------- Riverside (33): rivcoawm PUR XLSX ----------
def riverside():
    raw = fetch("https://rivcoawm.org/sites/g/files/aldnop221/files/2024-06/PUR-2024.xlsx")
    return _xlsx_permit_name(raw, "Permit #", "Permitee", "Riverside", "riverside-pur-2024")

# ---------- Santa Barbara (42): Box-hosted XLSX ----------
def santa_barbara():
    raw = fetch("https://cosantabarbara.app.box.com/index.php?rm=box_download_shared_file"
                "&shared_name=jdt95fy7gst3g8649l9t3ukrorr5xeh9&file_id=f_1821024891639")
    return _xlsx_permit_name(raw, "Permit Number", "Operator", "Santa Barbara",
                             "santabarbara-permits")

# ---------- ArcGIS-hosted county permit rosters (found 2026-07-01) ----------
def fresno():          # code 10
    return arcgis("https://services.arcgis.com/0xnwbwUttaTjns4i/arcgis/rest/services/"
                  "Fresno_County_permit_data/FeatureServer/0",
                  "Permit_Number", "Operator", "Fresno", "fresno-permit-data")

def san_diego():       # code 37 (MapServer layer 6)
    return arcgis("https://gis-public.sandiegocounty.gov/arcgis/rest/services/"
                  "AWM/AWM_Basemap/MapServer/6",
                  "permit_num", "permittee", "San Diego", "sandiego-awm-caps")

def napa():            # code 28
    return arcgis("https://gis.napacounty.gov/arcgis/rest/services/Hosted/"
                  "CalAgPermits_Field_Boundaries/FeatureServer/0",
                  "permit_number", "permittee", "Napa", "napa-calagpermits")

def colusa():          # code 06 (multi-county walnut map; also feeds Sutter etc.)
    return arcgis("https://services3.arcgis.com/zbiy4hH0vAQCfqtS/arcgis/rest/services/"
                  "WalnutMap/FeatureServer/0",
                  "permit__4", "permittee", "Colusa", "colusa-walnutmap")

def santa_cruz():      # code 44
    return arcgis("https://services1.arcgis.com/jJfZghspGKh8J9Jm/arcgis/rest/services/"
                  "Agricultural_Fields/FeatureServer/0",
                  "permit_num", "permittee", "Santa Cruz", "santacruz-ag-fields")

def yolo():            # code 57 (Crops_2024 layer 17)
    return arcgis("https://services2.arcgis.com/RETsakmE0SJfZXCd/arcgis/rest/services/"
                  "Crops_2024/FeatureServer/17",
                  "permit_num", "permittee", "Yolo", "yolo-crops-2024")

# ---------- county ArcGIS orgs (found 2026-07-01 via org enumeration) ----------
def merced():          # code 24
    return arcgis("https://services6.arcgis.com/LYh3hRvKq5ASgAVM/arcgis/rest/services/"
                  "Commodity/FeatureServer/0",
                  "PermNumber", "Operator", "Merced", "merced-commodity")

def kings():           # code 16
    return arcgis("https://services3.arcgis.com/24gLq1DBBzDfd0cZ/arcgis/rest/services/"
                  "Cotton_Plowdown_Map_WFL1/FeatureServer/0",
                  "PermNumber", "Operator", "Kings", "kings-permit-map")

def sutter():          # code 51 (on-prem ArcGIS Server MapServer; also carries some Yuba 58)
    return arcgis("https://gis.suttercounty.org/server/rest/services/TRAKiT/"
                  "TrakitCommDev/MapServer/43",
                  "permit_num", "permittee", "Sutter", "sutter-trakit-permits")

def main():
    dburl = os.environ.get("DBURL")
    if not dburl:
        sys.exit("Set DBURL in the environment.")
    for label, fn in [("plumas_local", plumas_local), ("monterey", monterey), ("kern", kern),
                      ("stanislaus", stanislaus), ("san_joaquin", san_joaquin),
                      ("contra_costa", contra_costa), ("riverside", riverside),
                      ("santa_barbara", santa_barbara), ("fresno", fresno),
                      ("san_diego", san_diego), ("napa", napa), ("colusa", colusa),
                      ("santa_cruz", santa_cruz), ("yolo", yolo), ("merced", merced),
                      ("kings", kings), ("sutter", sutter)]:
        try:
            c = fn(); print(f"  {label}: {c:,} source rows fetched")
        except Exception as e:
            print(f"  {label}: FAILED ({type(e).__name__}: {e})")
    print(f"Total distinct permit numbers with names: {len(rows):,}")
    if not rows:
        sys.exit("No name data fetched.")

    tmp = tempfile.NamedTemporaryFile("w", delete=False, newline="", suffix=".csv", encoding="utf-8")
    w = csv.writer(tmp)
    for p, (nm, cty, src) in rows.items():
        w.writerow([p, nm, cty, src])
    tmp.close()

    sql = f"""
create temp table _perm (permnum text, name text, county text, source text);
\\copy _perm from '{tmp.name.replace(os.sep,'/')}' csv
insert into public.operator_names (operator_id, name, entity_type, source, county, updated)
select a.owner, p.name, null, p.source, p.county, current_date::text
from (select distinct owner from public.applications
      where owner is not null and length(owner)>=7 and owner ~ '[0-9]') a
join _perm p on right(a.owner,7) = p.permnum
on conflict (operator_id) do update set
  name=excluded.name, source=excluded.source, county=excluded.county, updated=excluded.updated;
select count(*) as operator_ids_named from public.operator_names;
"""
    sqlf = tempfile.NamedTemporaryFile("w", delete=False, suffix=".sql", encoding="utf-8")
    sqlf.write("set statement_timeout=0;\n" + sql)
    sqlf.close()
    r = subprocess.run(["psql", dburl, "-v", "ON_ERROR_STOP=1", "-f", sqlf.name],
                       capture_output=True, text=True)
    os.unlink(tmp.name); os.unlink(sqlf.name)
    print(r.stdout.strip())
    if r.returncode != 0:
        sys.exit(r.stderr.strip())

if __name__ == "__main__":
    main()
